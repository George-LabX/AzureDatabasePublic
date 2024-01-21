# Databricks notebook source


# COMMAND ----------

import pandas as pd
import io
import datetime
import re
import numpy as np
from ast import literal_eval
import warnings
import os
from openpyxl import load_workbook
warnings.filterwarnings('ignore')

# COMMAND ----------

# MAGIC %md
# MAGIC ## Config

# COMMAND ----------

RFID_COC = pd.read_csv('', index_col=0)
df_existed = pd.read_csv('')
existed = df_existed['subject'].to_numpy()

characteristics_SHOCK = ['rfid', 'subject', 'room', 'cohort', 'trial_id', 'drug', 'box',
       'start_time', 'end_time', 'start_date', 'end_date',
       'total_active_lever_presses', 'total_inactive_lever_presses',
       'total_shocks', 'total_reward', 'rewards_after_first_shock',
       'rewards_got_shock', 'reward_timestamps']

# COMMAND ----------

# MAGIC %md
# MAGIC ## Helper Functions

# COMMAND ----------

# reformat shock id
def reformat_shock_id(shock_id, cohort):
    if 'PRESHOCK' in shock_id:
        return 'PRESHOCK'
    elif cohort in range(1,6):
        return 'SHOCK' + '_V' + str(int(shock_id[5:]))
    else:
        return 'SHOCK_V3'
    
# return valid list of datapoints
def process_datapoints(lst):
    while lst and lst[-1] == 0:
        lst.pop()
        
    if len(lst) == 0:
        return None
    else:
        return lst

# serialize timestamps
def serialize_timestamps(lst):
    if not lst:
        return None
    while lst[-1] == 0:
        lst.pop()
    return " ".join([str(i) for i in lst])

# get all the sheetnames in one excel workbook
def get_sheetnames_xlsx(file_name):
    wb = load_workbook(file_name, read_only=True, keep_links=False)
    return wb.sheetnames

# clean subject id
def clean_subject_id(sid):
    sid = sid.upper()
    if 'F' in sid:
        char = 'F'
    if 'M' in sid:
        char = 'M'

# COMMAND ----------

# MAGIC %md
# MAGIC ## Main Code

# COMMAND ----------

# MAGIC %md
# MAGIC ### General/Newer Data

# COMMAND ----------

def transform_shock(filepath):
    with open(filepath, "rb") as f:
        file_content = f.read()
    # Load the binary data into a pandas DataFrame
    df_raw = pd.read_excel(io.BytesIO(file_content), engine='openpyxl')

    # remove extra
    num_subjects = len(set([i for i in df_raw.iloc[5,:].values if isinstance(i, int)]))
    if df_raw.shape[1] > num_subjects+1:
        df_raw = df_raw.iloc[:,:num_subjects+1]

    df_raw = df_raw.T
    df_raw.reset_index(inplace=True)

    # modify the header
    new_header = df_raw.iloc[0]   #grab the first row for the header
    df = df_raw[1:]               #take the data except the header row
    df.columns = new_header 
    df.reset_index(drop=True, inplace=True)
    df.drop(['Filename', 'Experiment', 'Group', 'MSN', 'FR'], axis=1, inplace=True)

    # change data types
    cols = df.columns.tolist()
    int_columns = ['box','total shocks','total reward']

    for col in cols:
        name = col.lower()
        if ('active' in name) or ('reward' in name) or (name in int_columns):
            df[col] = df[col].astype('int32')
        elif ('date' in name):
            df[col] = df[col].apply(lambda x: datetime.datetime.strptime(x, "%Y-%m-%d").date())
        elif ('time' in name):
            df[col] = df[col].apply(lambda x: datetime.datetime.strptime(x, "%H:%M:%S").time())
        else:
            pass

    # reorganize the columns
    colnames = df.columns.tolist()

    reward_shock_begin = colnames.index('Reward # Got Shock 1')
    reward_col_begin = colnames.index('Reward 1')
    reward_col_end = colnames.index('Reward 201')

    df['Rewards Got Shock'] = df.iloc[:,reward_shock_begin:reward_col_begin].values.tolist()
    df['Rewards Got Shock'] = df['Rewards Got Shock'].apply(process_datapoints)
    df['Rewards Got Shock'] = df['Rewards Got Shock'].apply(serialize_timestamps)
    df['Reward Timestamps'] = df.iloc[:,reward_col_begin:reward_col_end+1].values.tolist()
    df['Reward Timestamps'] = df['Reward Timestamps'].apply(process_datapoints)
    df['Reward Timestamps'] = df['Reward Timestamps'].apply(serialize_timestamps)

    df.drop(df.iloc[:, reward_shock_begin:reward_col_end+1], inplace=True, axis=1)
    fname = f.name.split('/')[-1].split('.')[0]
    modified_filename = fname.replace('-','0')

    # parse the file name
    if modified_filename[0] == 'C':
        parser = r"(\AC[0-9]{2})HS((?:PRESHOCK[0-9]*|SHOCK[0-9]*))"
        cohort, shock_id = re.findall(parser, modified_filename)[0]
        room = None
    else:
        parser = r"(\A[A-Z]+[0-9]+[A-Z|0-9]{1})(C[0-9]{2})HS[COCAINE]*((?:PRESHOCK[0-9]*|SHOCK[0-9]*))"
        room, cohort, shock_id = re.findall(parser, modified_filename)[0]

    cohort = int(cohort[1:])
    trial_id = reformat_shock_id(shock_id, cohort)

    df['room'] = [room] * len(df)
    df['cohort'] = [cohort] * len(df)
    df['trial_id'] = [trial_id] * len(df)
    df['drug'] = ['cocaine'] * len(df)

    # get the final output
    df.rename(columns=str.lower,inplace=True)
    df = pd.merge(df, RFID_COC,  how='left', on = ['subject'])
    df.columns = df.columns.str.replace(' ','_')
    df.fillna({'rfid':-999}, inplace=True)
    df['rfid'] = df['rfid'].astype('int64')
    df = df[df['rfid'] != -999]
    df = df[characteristics_SHOCK]
    df = df.sort_values(by='box', ignore_index=True)

    output_path = ''
    df.to_csv(os.path.join(output_path, fname + '.csv'), index=False)

# COMMAND ----------

# MAGIC %md
# MAGIC ### SA / Old Data

# COMMAND ----------

def transform_old_shock(wb,ws):
    with open(wb, "rb") as f:
        file_content = f.read()
    df_raw = pd.read_excel(io.BytesIO(file_content), engine='openpyxl', sheet_name = ws).T.reset_index()

    # modify the header
    new_header = df_raw.iloc[0]     #grab the first row for the header
    df = df_raw[1:]                 #take the data except the header row
    df.columns = new_header 

    # get rid of 0s
    df.replace(0, np.nan, inplace=True)
    df.dropna(how='all', axis=1, inplace=True)
    df.fillna(0,inplace=True)

    # group timestamp columns
    colnames = df.columns.tolist()
    reward_shock_begin = colnames.index('Reward # Got Shock 1')
    reward_col_begin = colnames.index('Reward 1')
    reward_col_end = colnames.index('Rewards After First Shock')
    df['Rewards Got Shock'] = df.iloc[:,reward_shock_begin:reward_col_begin].values.tolist()
    df['Rewards Got Shock'] = df['Rewards Got Shock'].apply(process_datapoints)
    df['Rewards Got Shock'] = df['Rewards Got Shock'].apply(serialize_timestamps)
    df['Reward Timestamps'] = df.iloc[:,reward_col_begin:reward_col_end].values.tolist()
    df['Reward Timestamps'] = df['Reward Timestamps'].apply(process_datapoints)
    df['Reward Timestamps'] = df['Reward Timestamps'].apply(serialize_timestamps)
    df.drop(df.iloc[:, reward_shock_begin:reward_col_end], inplace=True, axis=1)

    # add extra info
    cohort = int(wb.split('/')[-1][1:3])
    raw_trial_id = ws.split('_')[0]

    if raw_trial_id == 'PRESHOCK':
        trial_id = raw_trial_id
    elif cohort in range(1,6):
        trial_id = raw_trial_id.replace('0','_V')
    else:
        trial_id = 'SHOCK_V3'

    df['room'] = [None] * len(df)
    df['cohort'] = [cohort] * len(df)
    df['trial_id'] = [trial_id] * len(df)
    df['drug'] = ['cocaine'] * len(df)
    df['end_time'] = [datetime.datetime.min.time()] * len(df)
    df['end_date'] = [datetime.datetime.min.date()] * len(df)

    # reorganize columns
    df.rename(columns=str.lower,inplace=True)
    df.columns = df.columns.str.replace(' ','_')
    dff = pd.merge(df, RFID_COC,  how='left', on = ['subject'])
    dff.fillna({'rfid':-999}, inplace=True)
    dff['rfid'] = dff['rfid'].astype('int64')
    dff = dff[characteristics_SHOCK]
    dff['start_time'] = dff['start_time'].apply(lambda x: datetime.datetime.strptime(x, "%H:%M:%S").time())
    dff['start_date'] = dff['start_date'].apply(lambda x: datetime.datetime.strptime(x, "%m/%d/%Y").date())
    dff.sort_values(by='subject', inplace=True)
    print(len(dff))
    dff = dff[~dff['subject'].isin(existed)]
    print(len(dff))
    dff['rfid'] = dff['rfid'].astype('int64')
    dff = dff[dff['rfid'] != -999]
    filename = wb.split('/')[-1][:3] + '_' + ws.split('.')[0]+'.csv' 

    output_path = ''
    dff.to_csv(os.path.join(output_path, filename), index=False)

# COMMAND ----------

# MAGIC %md
# MAGIC ## Run Code

# COMMAND ----------

for folder in dbutils.fs.ls(''):
    if 'SHOCK' in folder.path:
        folder_path = folder.path
        for f in dbutils.fs.ls(folder_path):
            filepath = "/" + f.path.replace(':','')
            if 'sa' in filepath:
                wb = filepath
                worksheets = sorted(get_sheetnames_xlsx(wb))
                for ws in worksheets:
                    print(wb,ws)
                    transform_old_shock(wb,ws)
            else:
                if ('Backup' not in filepath) and ('SHOCK' in filepath):
                    print(filepath)
                    transform_shock(filepath)
