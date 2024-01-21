# Databricks notebook source


# COMMAND ----------

import pandas as pd
import io
import datetime
import re
import numpy as np
from ast import literal_eval
import warnings
import os
from openpyxl import load_workbook
warnings.filterwarnings('ignore')

# COMMAND ----------

# MAGIC %md
# MAGIC ## Config

# COMMAND ----------

RFID_OXY = pd.read_csv('', index_col=0)
RFID_COC = pd.read_csv('', index_col=0)

characteristics_PR = ['rfid', 'subject', 'room', 'cohort', 'trial_id', 'drug', 'box','start_time', 'end_time', 
 'start_date', 'end_date', 'breakpoint', 'last_ratio', 'ratios', 'active_lever_presses', 'inactive_lever_presses',
 'reward_presses']

coc_rewards = np.arange(19)
coc_lr = [0,1,2,4,6,9,12,15,20,25,32,40,50,62,77,95,118,145,178]
coc_lr_dict = dict(zip(coc_rewards,coc_lr))

oxy_rewards = np.arange(69)
oxy_lr = [0,1,1,2,2,3,3,4,4,5,5,6,6,7,7,8,8,9,9,10] + np.arange(10,49).tolist() + [50,60,70,80,90,100,100,100,100,100]
oxy_lr_dict = dict(zip(oxy_rewards,oxy_lr))

# COMMAND ----------

updates = pd.read_csv('/dbfs/mnt/testmount/data/update_filelist_pr.csv')
UPDATE_LIST = list(updates['files'])

# COMMAND ----------

# MAGIC %md
# MAGIC ## Helper Functions

# COMMAND ----------

# retrieve the breakpoint
def get_last_ratio(lr_dict, breakpoint):
    if not breakpoint or np.isnan(breakpoint):
        return None
    if breakpoint not in lr_dict:
        return None
    lr_list = list(lr_dict.values())
    idx = lr_list.index(breakpoint) + 1
    return lr_list[idx]

# return valid list of datapoints
def process_datapoints(lst):
    while lst and lst[-1] == 0:
        lst.pop()
        
    if len(lst) == 0:
        return None
    else:
        return lst

# standardize trial id
def process_trial_id(tid):
    i = 0
    while not (tid[i].isdigit()):
        i += 1
    name,num = tid[:i],tid[i:]
    res = name + num.rjust(2, "0")
    return res

# serialize timestamps
def serialize_timestamps(lst):
    if not lst:
        return None
    while lst[-1] == 0:
        lst.pop()
    return " ".join([str(i) for i in lst])

# get all the sheetnames in one excel workbook
def get_sheetnames_xlsx(file_name):
    wb = load_workbook(file_name, read_only=True, keep_links=False)
    return wb.sheetnames

# clean subject id
def clean_subject_id(sid):
    sid = sid.upper()
    if 'F' in sid:
        char = 'F'
    if 'M' in sid:
        char = 'M'
        
    idx = sid.index(char)
    return sid[idx:].split('.')[0]

# convert column names into correct format
def clean_cols(s):
    if 'Y' in s:
        return s.replace('Y','Active ')
    elif 'U' in s:
        return s.replace('U','Inactive ')
    elif 'V' in s:
        return s.replace('V','Reward ')
    else:
        return s

# COMMAND ----------

# MAGIC %md
# MAGIC ## Main Code

# COMMAND ----------

# MAGIC %md
# MAGIC ### General/Newer Data

# COMMAND ----------

def transform_pr(filepath):
    # import data and transpose
    #print(filepath)
    with open(filepath, "rb") as f:
        file_content = f.read()

    output_path = '/dbfs/mnt/testmount/output/PR/'
    if os.path.exists(os.path.join(output_path, f.name.split('/')[-1].split('.')[0] + '.csv')) and f.name.split('/')[-1] not in UPDATE_LIST:
        print(f.name.split('/')[-1]+': skipped')
        return
    print(filepath)

    # Load the binary data into a pandas DataFrame
    df_raw = pd.read_excel(io.BytesIO(file_content), engine='openpyxl')
    num_subjects = len(set([i for i in df_raw.iloc[5,:].values if isinstance(i, int)]))

    if df_raw.shape[1] > num_subjects+1:
        df_raw = df_raw.iloc[:,:num_subjects+1]

    df_raw = df_raw.T
    df_raw.reset_index(inplace=True)

    # modify the header
    new_header = df_raw.iloc[0]   #grab the first row for the header
    df = df_raw[1:]               #take the data except the header row
    df.columns = new_header 
    df.reset_index(drop=True, inplace=True)
    df.drop(['Filename', 'Experiment', 'Group', 'MSN', 'FR'], axis=1, inplace=True)

    # change data types
    cols = df.columns.tolist()
    int_columns = ['box','last ratio']

    for col in cols:
        name = col.lower()
        test_item = df[col][0]
        if ('active' in name) or ('reward' in name) or (name in int_columns):
            df[col] = df[col].apply(lambda x: int(x) if not pd.isnull(x) else None)
        elif (("timestamps" in name) or (name in ['ratios', 'rewards_got_shock'])): 
            df[col] = df[col].apply(lambda x: serialize_timestamps(x))
        elif ('date' in name):
            if not isinstance(test_item, datetime.date):
                df[col] = df[col].apply(lambda x: datetime.datetime.strptime(x, "%Y-%m-%d").date() if not pd.isnull(x) else None)
        elif ('time' in name and 'timeout' not in name):
            if not isinstance(test_item, datetime.time):
                df[col] = df[col].apply(lambda x: datetime.datetime.strptime(x, "%H:%M:%S").time() if not pd.isnull(x) else None)
        else:
            pass

    # reorganize the columns
    colnames = df.columns.tolist()
    reward_col_begin = colnames.index('Reward 1')
    df['ratios'] = df.iloc[:, reward_col_begin:].values.tolist()
    df['ratios'] = df['ratios'].apply(process_datapoints)
    df['ratios'] = df['ratios'].apply(serialize_timestamps)
    points_col_begin = df.columns.tolist().index('ratios')
    df.drop(df.iloc[:, reward_col_begin:points_col_begin], inplace=True, axis=1)
    df.rename(columns={"Reward": "Reward Presses"}, inplace=True)

    # parse the file name
    file = filepath.split('/')[-1]
    parsers = [r"(\A[A-Z]+[0-9]+[A-Z|0-9]{1})(C[0-9]{2})[HS]*[COCAINE|OXY]*((?:LGA|SHA|PR|TREATMENT)[0-9]+)_output",
               r"(\AC[0-9]{2})[HS]*[OXY]*((?:LGA|SHA|PR|TREATMENT)[0-9]+)_output"]
    if file[0] == 'C':
        parser = parsers[1]
        cohort, trial_id = re.findall(parser, file)[0]
        room = None
    else:
        parser = parsers[0]
        room, cohort, trial_id = re.findall(parser, file)[0]

    df['room'] = [room] * len(df)
    cohort = int(cohort[1:])
    df['cohort'] = [cohort] * len(df)
    trial_id = process_trial_id(trial_id)
    df['trial_id'] = [trial_id] * len(df)
    
    # merge in the RFID and reorganize the column formats
    df.rename(columns=str.lower,inplace=True)
    df.columns = df.columns.str.replace(' ','_')
    if 'coc' in filepath:
        rfid_to_merge = RFID_COC
        lr_dict = coc_lr_dict
        df['drug'] = ['cocaine'] * len(df)
    if 'oxy' in filepath:
        rfid_to_merge = RFID_OXY
        lr_dict = oxy_lr_dict
        df['drug'] = ['oxycodone'] * len(df)


    # calculate special variables 
    df['breakpoint'] = df['reward_presses'].apply(lambda x: lr_dict[x] if x in lr_dict else None)
    df['last_ratio'] = df['breakpoint'].apply(lambda x: get_last_ratio(lr_dict, x))

    df = pd.merge(df, rfid_to_merge,  how='left', on = ['subject'])
    df.columns = df.columns.str.replace(' ','_')
    df.fillna({'rfid':-999}, inplace=True)
    df['rfid'] = df['rfid'].astype('int64')
    df = df[df['rfid'] != -999]
    df = df[characteristics_PR]
    df = df.sort_values(by='box')
    df.drop_duplicates(inplace=True)
    df = df.reset_index(drop=True)

    if len(set(df.subject)) < len(df.subject):
        print(filepath)

    df.to_csv(os.path.join(output_path, f.name.split('/')[-1].split('.')[0] + '.csv'), index=False)

# COMMAND ----------

# MAGIC %md
# MAGIC ### SA / Old Data

# COMMAND ----------

def transform_old_pr(wb, ws):
    #print(wb, ws)
    with open(wb, "rb") as f:
        file_content = f.read()

    filename = ws.split('.')[0]+'_transformed.csv'
    output_path = '/dbfs/mnt/testmount/output/PR/'
    if os.path.exists(os.path.join(output_path, filename)):
        print(wb+" : "+ws+' skipped')
        return
    print(wb+" : "+ws)

    # Load the binary data into a pandas DataFrame
    df_raw = pd.read_excel(io.BytesIO(file_content), engine='openpyxl', sheet_name=ws).T.reset_index()

    # modify the header
    new_header = df_raw.iloc[0]     #grab the first row for the header
    df = df_raw[1:]                 #take the data except the header row
    df.columns = new_header 

    # get rid of 0s
    df.replace(0, np.nan, inplace=True)
    df.dropna(how='all', axis=1, inplace=True)
    df.fillna(0,inplace=True)
    
    # transform columns names
    ID_col = df.columns.tolist()[0]
    filtered_cols = [i for i in df.columns if i[0] in ['V'] or 
                     i in [ID_col, 'Active Lever Presses', 'Inactive Lever Presses', 'Reward']]
    dff = df[filtered_cols]
    new_cols = [clean_cols(i) for i in dff.columns]
    dff.columns = new_cols
    
    # add extra info
    if 'OXY' in ws:
        drug = 'oxycodone'
        parser = r'(\AC[0-9]{2})HS[OXY]*((?:PR|TREATMENT)[0-9]+)'
        rfid_to_merge = RFID_OXY
        lr_dict = oxy_lr_dict
    else:
        drug = 'cocaine'
        parser = r'(C[0-9]{2})HS((?:PR|TREATMENT)[0-9]{2})'
        rfid_to_merge = RFID_COC
        lr_dict = coc_lr_dict

    if '-' in ws:
            to_split = '-'
    if '_' in ws:
        to_split = '_'
        
    info, date = ws.split('.')[0].split(to_split)
    cohort, trial_id = re.findall(parser, ws)[0]
    dt = pd.to_datetime(date, format='%Y%m%d', errors='ignore')
    
    dff['room'] = [None] * len(dff)
    dff['cohort'] = [cohort[1:]] * len(dff)
    trial_id = process_trial_id(trial_id)
    dff['trial_id'] = [trial_id] * len(dff)
    dff['drug'] = [drug] * len(dff)
    dff['box'] = [None] * len(dff)
    dff['start_time'] = [datetime.datetime.min.time()] * len(dff)
    dff['end_time'] = [datetime.datetime.min.time()] * len(dff)
    dff['start_date'] = [dt] * len(dff)
    dff['end_date'] = [datetime.datetime.min.date()] * len(dff)
    
    # reorganize the columns
    colnames = dff.columns.tolist()
    reward_col_begin = colnames.index('Reward 0')
    reward_col_end = colnames.index('room')
    dff['ratios'] = [None] * len(dff)
    dff.drop(dff.iloc[:, reward_col_begin:reward_col_end], inplace=True, axis=1)
    dff.rename(columns={"Reward": "Reward Presses", ID_col:"subject"}, inplace=True)
    
    # calculate special variables 
    dff['breakpoint'] = dff['Reward Presses'].apply(lambda x: lr_dict[x] if x in lr_dict else None)
    dff['last_ratio'] = dff['breakpoint'].apply(lambda x: get_last_ratio(lr_dict, x))
    
    dff = pd.merge(dff, rfid_to_merge, how='left', on = ['subject'])
    dff.rename(columns=str.lower,inplace=True)
    dff.columns = dff.columns.str.replace(' ','_')
    dff.fillna({'rfid':-999}, inplace=True)
    dff['rfid'] = dff['rfid'].astype('int64')
    dff = dff[dff['rfid'] != -999]
    dff = dff[characteristics_PR]
    
    dff.to_csv(os.path.join(output_path, filename), index=False)

# COMMAND ----------

# MAGIC %md
# MAGIC ## Run code

# COMMAND ----------

for folder in dbutils.fs.ls(''):
    if ('PR_general' in folder.path) or ('OLD_SA' in folder.path):
        folder_path = folder.path
        for f in dbutils.fs.ls(folder_path):
            filepath = "/" + f.path.replace(':','')
            if 'OLD_SA' in filepath:
                wb = filepath
                worksheets = sorted(get_sheetnames_xlsx(wb))
                worksheets = [ws for ws in worksheets if 'PR' in ws or 'TREATMENT' in ws]
                for ws in worksheets:
                    transform_old_pr(wb,ws)
            else:
                if ('DISSECT' not in filepath) and ('TEST' not in filepath) and ('PRETREAT' not in filepath):
                    transform_pr(filepath)

# COMMAND ----------


