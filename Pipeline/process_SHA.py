# Databricks notebook source


# COMMAND ----------

import pandas as pd
import io
import datetime
import re
import numpy as np
from ast import literal_eval
import warnings
import os
from openpyxl import load_workbook
warnings.filterwarnings('ignore')

# COMMAND ----------

# MAGIC %md
# MAGIC ## Config

# COMMAND ----------

updates = pd.read_csv('')
UPDATE_LIST = list(updates['files'])
UPDATE_LIST

# COMMAND ----------

RFID_OXY = pd.read_csv('', index_col=0)
RFID_COC = pd.read_csv('', index_col=0)

characteristics_LGA_SHA = ['rfid','subject','room','cohort','trial_id','drug','box', 'start_time', 'end_time',
'start_date','end_date','active_lever_presses','inactive_lever_presses','reward_presses','timeout_presses',
'active_timestamps','inactive_timestamps','reward_timestamps','timeout_timestamps']

# COMMAND ----------

# MAGIC %md
# MAGIC ### Helper Functions

# COMMAND ----------

# get all the sheetnames in one excel workbook
def get_sheetnames_xlsx(file_name):
    wb = load_workbook(file_name, read_only=True, keep_links=False)
    return wb.sheetnames

# clean subject id
def clean_subject_id(sid):
    sid = sid.upper()
    if 'F' in sid:
        char = 'F'
    if 'M' in sid:
        char = 'M'
        
    idx = sid.index(char)
    return sid[idx:].split('.')[0]

# convert column names into correct format
def clean_cols(s):
    if 'Y' in s:
        return s.replace('Y','Active ')
    elif 'U' in s:
        return s.replace('U','Inactive ')
    elif 'V' in s:
        return s.replace('V','Reward ')
    else:
        return s
    
# return valid list of datapoints
def process_datapoints(lst):
    while lst and lst[-1] == 0:
        lst.pop()
        
    if len(lst) == 0:
        return None
    else:
        return lst

# count valid data points
def count_datapoints(lst):
    while lst and lst[-1] == 0:
        lst.pop()
        
    if len(lst) == 0:
        return None
    else:
        return len(lst)

# serialize timestamps
def serialize_timestamps(lst):
    if not lst:
        return None
    while lst[-1] == 0:
        lst.pop()
    return " ".join([str(i) for i in lst])

# COMMAND ----------

# MAGIC %md
# MAGIC ## Main Code

# COMMAND ----------

# MAGIC %md
# MAGIC ### General/Newer Data

# COMMAND ----------

def transform_lga_sha(filepath):
    with open(filepath, "rb") as f:
        file_content = f.read()

    fname = f.name.split('/')[-1].split('.')[0].upper()
    output_path = ''
    if os.path.exists(os.path.join(output_path, fname + '.csv')) and f.name.split('/')[-1] not in UPDATE_LIST:
        return
    print(f.name)
    # Load the binary data into a pandas DataFrame
    df_raw = pd.read_excel(io.BytesIO(file_content), engine='openpyxl').T
    df_raw.reset_index(inplace=True)

    # modify the header
    new_header = df_raw.iloc[0]   #grab the first row for the header
    df = df_raw[1:]               #take the data except the header row
    df.columns = new_header 
    df.reset_index(drop=True, inplace=True)
    df.drop(['Filename', 'Experiment', 'Group', 'MSN', 'FR'], axis=1, inplace=True)
    df.drop_duplicates(inplace=True)

    # patch
    if ('Timeout Press 1') not in df.columns:
        df['Timeout Press 1'] = [0] * len(df)
        
    # change data types
    cols = df.columns.tolist()
    for col in cols:
        name = col.lower()
        if ('active' in name) or ('reward' in name) or ('timeout' in name) or (name == 'box'):
            df[col] = df[col].apply(lambda x: int(x) if not pd.isnull(x) else x)
        elif ('date' in name):
            df[col] = df[col].apply(lambda x: pd.to_datetime(x, format='%Y-%m-%d', errors='ignore'))
        elif ('time' in name):
            df[col] = df[col].apply(lambda x: datetime.datetime.strptime(x, "%H:%M:%S").time() if type(x)==str else datetime.datetime.strptime('00:00:00', "%H:%M:%S").time())
        else:
            pass
        
    # group the timestamps
    colnames = df.columns.tolist()
    active_col_begin = colnames.index('Active 1')
    inactive_col_begin = colnames.index('Inactive 1')
    reward_col_begin = colnames.index('Reward 1')
    timeout_col_begin = colnames.index('Timeout Press 1')
    idx_end = df.shape[1]
    df['Active Timestamps'] = df.iloc[:, active_col_begin:inactive_col_begin].values.tolist()
    df['Active Timestamps'] = df['Active Timestamps'].apply(process_datapoints)
    df['Active Timestamps'] = df['Active Timestamps'].apply(serialize_timestamps)
    df['Inactive Timestamps'] = df.iloc[:, inactive_col_begin:reward_col_begin].values.tolist()
    df['Inactive Timestamps'] = df['Inactive Timestamps'].apply(process_datapoints)
    df['Inactive Timestamps'] = df['Inactive Timestamps'].apply(serialize_timestamps)
    df['Reward Timestamps'] = df.iloc[:, reward_col_begin:timeout_col_begin].values.tolist()
    df['Reward Timestamps'] = df['Reward Timestamps'].apply(process_datapoints)
    df['Reward Timestamps'] = df['Reward Timestamps'].apply(serialize_timestamps)
    df['Timeout Timestamps'] = df.iloc[:, timeout_col_begin:idx_end].values.tolist()
    df['Timeout Timestamps'] = df['Timeout Timestamps'].apply(process_datapoints)
    df['Timeout Timestamps'] = df['Timeout Timestamps'].apply(serialize_timestamps)

    # reorganize the columns
    timestamp_col_begin = df.columns.tolist().index('Active Timestamps')
    df.drop(df.iloc[:, active_col_begin:timestamp_col_begin], inplace=True, axis=1)
    df.rename(columns={"Reward": "Reward Presses"}, inplace=True)
    df['Timeout Presses'] = df['Active Lever Presses'] - df['Reward Presses']

    # parse the filename
    if fname[0] == 'C':
        parser = r"(\AC[0-9]{2})[HS]*[OXY]*((?:LGA|SHA)[0-9]{2})"
        cohort, trial_id = re.findall(parser, fname)[0]
        room = None
    else:
        parser = r"(\A[A-Z]+[0-9]+[A-Z|0-9]{1})(C[0-9]{2})[HS|4S]*[OXY|COC|COCAINE]*((?:LGA|SHA)[0-9]{2})"
        room, cohort, trial_id = re.findall(parser, fname)[0]

    df['room'] = [room] * len(df)
    cohort = int(cohort[1:])
    df['cohort'] = [cohort] * len(df)
    df['trial_id'] = [trial_id] * len(df)

    if 'oxy' in fname.lower():
        df['drug'] = ['oxycodone'] * len(df)
        rfid_to_merge = RFID_OXY
    else:
        df['drug'] = ['cocaine'] * len(df)
        rfid_to_merge = RFID_COC

    # merge in the RFID and reorganize the column formats
    df.rename(columns=str.lower,inplace=True)
    df = pd.merge(df, rfid_to_merge,  how='left', on = ['subject'])
    df.columns = df.columns.str.replace(' ','_')
    df.fillna({'rfid':-999}, inplace=True)
    df['rfid'] = df['rfid'].astype('int64')
    df = df[df['rfid'] != -999]
    df.drop_duplicates(inplace=True)
    df = df[characteristics_LGA_SHA]
    
    # if len(set(df.subject)) < len(df.subject):
    #     print(filepath)

    output_path = ''
    df.to_csv(os.path.join(output_path, fname + '.csv'), index=False)

# COMMAND ----------

# MAGIC %md
# MAGIC ### SA/Old Data

# COMMAND ----------

def transform_old_lga_sha(wb, ws):
    with open(wb, "rb") as f:
        file_content = f.read()

    filename = ws.split('.')[0]
    output_path = ''
    if os.path.exists(os.path.join(output_path, filename + '.csv')):
        return
    print(wb+" : "+ws)
    # Load the binary data into a pandas DataFrame
    df_raw = pd.read_excel(io.BytesIO(file_content), engine='openpyxl', sheet_name = ws).T.reset_index()

    # modify the header
    new_header = df_raw.iloc[0]     #grab the first row for the header
    df = df_raw[1:]                 #take the data except the header row
    df.columns = new_header 

    # clean subject id
    ID_col = df.columns.tolist()[0]
    df[ID_col] = df[ID_col].apply(clean_subject_id)

    # get rid of 0s
    df.replace(0, np.nan, inplace=True)
    df.dropna(how='all', axis=1, inplace=True)
    df.fillna(0,inplace=True)

    # transform columns names
    filtered_cols = [i for i in df.columns if i[0] in ['U','V','Y','T'] or 
                     i in [ID_col, 'Active Lever Presses', 'Inactive Lever Presses', 'Reward']]
    dff = df[filtered_cols]
    new_cols = [clean_cols(i) for i in dff.columns]
    dff.columns = new_cols

    # add extra info
    if 'OXY' in ws:
        drug = 'oxycodone'
        parser = r'(C[0-9]{2})HSOXY((?:LGA|SHA)[0-9]{2})'
        rfid_to_merge = RFID_OXY
    else:
        drug = 'cocaine'
        parser = r'(C[0-9]{2})HS((?:LGA|SHA)[0-9]{2})'
        rfid_to_merge = RFID_COC

    if '-' in ws:
        to_split = '-'
    if '_' in ws:
        to_split = '_'

    info, date = ws.split('.')[0].split(to_split)
    cohort, trial_id = re.findall(parser, ws)[0]
    dt = pd.to_datetime(date, format='%Y%m%d', errors='ignore')

    dff['room'] = [None] * len(dff)
    dff['cohort'] = [cohort[1:]] * len(dff)
    dff['trial_id'] = [trial_id] * len(dff)
    dff['drug'] = [drug] * len(dff)
    dff['box'] = [None] * len(dff)
    dff['start_time'] = [datetime.datetime.min.time()] * len(dff)
    dff['end_time'] = [datetime.datetime.min.time()] * len(dff)
    dff['start_date'] = [dt] * len(dff)
    dff['end_date'] = [datetime.datetime.min.date()] * len(dff)

    # group the timestamps
    colnames = dff.columns.tolist()
    inactive_col_begin = colnames.index('Inactive 0')
    reward_col_begin = colnames.index('Reward 0')
    active_col_begin = colnames.index('Active 0')
    timeout_col_begin = colnames.index('Timeout Press 1')
    timeout_col_end = colnames.index('room')

    dff['Inactive Timestamps'] = dff.iloc[:, inactive_col_begin:reward_col_begin].values.tolist()
    dff['Inactive Timestamps'] = dff['Inactive Timestamps'].apply(process_datapoints)
    dff['Inactive Timestamps'] = dff['Inactive Timestamps'].apply(serialize_timestamps)
    dff['Reward Timestamps'] = dff.iloc[:, reward_col_begin:active_col_begin].values.tolist()
    dff['Reward Timestamps'] = dff['Reward Timestamps'].apply(process_datapoints)
    dff['Reward Timestamps'] = dff['Reward Timestamps'].apply(serialize_timestamps)
    dff['Active Timestamps'] = dff.iloc[:, active_col_begin:timeout_col_begin].values.tolist()
    dff['Active Timestamps'] = dff['Active Timestamps'].apply(process_datapoints)
    dff['Active Timestamps'] = dff['Active Timestamps'].apply(serialize_timestamps)
    dff['Timeout Timestamps'] = dff.iloc[:, timeout_col_begin:timeout_col_end].values.tolist()
    dff['Timeout Timestamps'] = dff['Timeout Timestamps'].apply(process_datapoints)
    dff['Timeout Presses'] = dff['Timeout Timestamps'].apply(lambda x: len(x) if x else None)
    dff['Timeout Timestamps'] = dff['Timeout Timestamps'].apply(serialize_timestamps)

    # reformat columns, merge rfid
    dff.drop(dff.iloc[:, inactive_col_begin:timeout_col_end], inplace=True, axis=1)
    dff.rename(columns={"Reward": "Reward Presses", ID_col:"subject"}, inplace=True)
    dff = pd.merge(dff, rfid_to_merge,  how='left', on = ['subject'])
    dff.rename(columns=str.lower,inplace=True)
    dff.columns = dff.columns.str.replace(' ','_')
    dff.fillna({'rfid':-999}, inplace=True)

    dff['active_lever_presses'] = dff['active_lever_presses'].astype('int64')
    dff['rfid'] = dff['rfid'].astype('int64')
    dff['rfid'] = dff['rfid'].astype('int64')
    dff['rfid'] = dff['rfid'].astype('int64')
    dff['rfid'] = dff['rfid'].astype('int64')
    dff = dff[characteristics_LGA_SHA]
    dff.sort_values(by='subject',inplace=True)
    dff = dff[dff['rfid'] != -999]
    
    dff.to_csv(os.path.join(output_path, filename + '.csv'), index=False)

# COMMAND ----------

# MAGIC %md
# MAGIC ## Run Code

# COMMAND ----------

for folder in dbutils.fs.ls(''):
    if ('SHA_general' in folder.path) or ('OLD_SA' in folder.path):
        folder_path = folder.path
        for f in dbutils.fs.ls(folder_path):
            filepath = "/" + f.path.replace(':','')
            if 'OLD_SA' in filepath:
                wb = filepath
                worksheets = sorted(get_sheetnames_xlsx(wb))
                worksheets = [ws for ws in worksheets if 'SHA' in ws]
                for ws in worksheets:
                    transform_old_lga_sha(wb, ws)
            else:
                if ('DISSECT' not in filepath) and ('PRETREATMENT' not in filepath) and ('Backup' not in filepath):
                    transform_lga_sha(filepath)

# COMMAND ----------


